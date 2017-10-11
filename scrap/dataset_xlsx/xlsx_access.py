import openpyxl as op
from openpyxl import load_workbook
from openpyxl import workbook

workbook = load_workbook(filename='Ramamurthy Data Set.xlsx')
print workbook.get_sheet_names()
sheet_ranges = workbook['Sheet'] #name of the worksheet
#print(sheet_ranges['A1'].value)
#print(sheet_ranges['A2'].value)
#print(sheet_ranges['A3'].value)
#print(sheet_ranges['A4'].value)

#print "Column 3:  " + (sheet_ranges['C1'].value)

#worksheet.write_cols(worksheet)
#worksheet.cell('cell').value


# string loop here
#cell = 'C2'
#for x in cell:
    #it = int((cell[1]))
    #it += 1
    #print "it", it
    #it=cell[1]
#print cell



#ws = workbook.get_sheet_by_name(name = 'Sheet1') 
#for row in ws.iter_rows('C1:C92'):
#    for cell in row:
#        print cell