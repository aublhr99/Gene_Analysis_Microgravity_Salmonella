#from openpyxl import load_workbook
#wb = load_workbook('Ramamurthy Data Set.xlsx.xlsx') 
#ws = wb.active
#for i in range(1,92):
#    for j in range(1,3):
#        print ws.cell(row = i, column = j).value

from openpyxl import load_workbook
wb = load_workbook(filename = 'Ramamurthy Data Set.xlsx.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)